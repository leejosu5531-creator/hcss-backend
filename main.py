# -*- coding: utf-8 -*-
import os
import io
import re
import json
import base64
import logging
import datetime as dt
from fastapi import FastAPI, HTTPException
from fastapi.middleware.cors import CORSMiddleware
from pydantic import BaseModel
from google.oauth2.service_account import Credentials
from googleapiclient.discovery import build
from googleapiclient.http import MediaIoBaseDownload
from openpyxl import load_workbook

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger("hcss_backend")

app = FastAPI(title="화원교회 찬양대 좌석 배치 API")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

SCOPES = ["https://www.googleapis.com/auth/drive.readonly"]
DRIVE_FILE_NAME = "26년 할렐루야 출석부.xlsx"

BASE_SEAT_ORDER = {
    1: ["남현숙", "김경미", "백시원", "박지현", "이서연", "이은진", "고태옥", "황진서", "장영자", "김은희", "황재연", "박유림",
        "임은애", "김혜진", "노인숙", "유미영", "박소윤", "오정민", "서정영", "박혜숙", "심윤정"],
    2: ["김은영", "정진순", "조한주", "손국희", "김서영", "김은희", "박진희", "황주경", "김은현", "박도희", "윤선미",
        "장형미", "조경화", "권순예", "김현경", "박미선", "강미화", "김준희", "방민주", "김지영", "우다연"],
    3: ["류은채", "노희령", "김지은", "강종훈", "권재훈", "하효동", "권영훈", "김기형", 
        "이입교", "권상대b", "성준호", "전병대", "목현성", "정동근", "이재관"],
    4: ["시진규", "남명호", "이병륜", "김경남", "김동근", "문지민", "박상훈", "권상대a", 
        "이종성", "박동민", "박선일", "장현재", "최한열", "이의헌", "김기훈", "김대성"]
}

ALTO_OUT_PRIORITY = ["박소윤", "오정민", "박혜숙", "심윤정", "박미선", "강미화", "유미영", "배수련"]
SOPRANO_ROW3_PRIORITY = ["류은채", "노희령", "김지은"]

def get_google_credentials():
    b64_str = os.environ.get("SERVICE_ACCOUNT_BASE64", "").strip()
    if not b64_str:
        raise ValueError("SERVICE_ACCOUNT_BASE64 환경변수가 설정되지 않았습니다.")
    decoded_json = base64.b64decode(b64_str).decode("utf-8")
    return Credentials.from_service_account_info(json.loads(decoded_json), scopes=SCOPES)

def google_drive_service():
    return build("drive", "v3", credentials=get_google_credentials())

def find_drive_file(service):
    q = f"name = '{DRIVE_FILE_NAME.replace('\'', '\\\'')}' and trashed = false"
    r = service.files().list(q=q, spaces="drive", fields="files(id,name)").execute()
    fs = r.get("files", [])
    if not fs:
        q_old = "name = '할렐루야 출석부.xlsx' and trashed = false"
        r_old = service.files().list(q=q_old, spaces="drive", fields="files(id,name)").execute()
        fs_old = r_old.get("files", [])
        if not fs_old:
            raise FileNotFoundError(f"구글 드라이브에서 '{DRIVE_FILE_NAME}' 파일을 찾지 못했습니다.")
        return fs_old[0]
    return fs[0]

def download_excel_to_stream(service, file_id):
    request = service.files().get_media(fileId=file_id)
    file_stream = io.BytesIO()
    downloader = MediaIoBaseDownload(file_stream, request)
    done = False
    while not done:
        _, done = downloader.next_chunk()
    file_stream.seek(0)
    return file_stream

def parse_cell_date(val, target_year):
    if val is None: return None
    if isinstance(val, dt.datetime): return val.date()
    if isinstance(val, dt.date): return val
    if isinstance(val, (int, float)):
        try: return (dt.datetime(1899, 12, 30) + dt.timedelta(days=val)).date()
        except: pass
    val_str = str(val).strip()
    if not val_str: return None
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d", "%Y.%m.%d", "%Y/%m/%d", "%Y%m%d"):
        try: return dt.datetime.strptime(val_str, fmt).date()
        except ValueError: pass
    match = re.search(r"(\d{1,2})[\/\.\-\s월]+(\d{1,2})", val_str)
    if match:
        try: return dt.date(target_year, int(match.group(1)), int(match.group(2)))
        except ValueError: pass
    return None

def read_attendance(file_stream, target_date):
    wb = load_workbook(file_stream, data_only=True)
    sheet_map = {"소프라노": "S", "알토": "A", "테너": "T", "베이스": "B"}
    people = {p: [] for p in "SATB"}
    target_year = target_date.year

    for sheet_name, part_code in sheet_map.items():
        if sheet_name not in wb.sheetnames: continue
        ws = wb[sheet_name]
        date_col = None
        for col in range(1, ws.max_column + 1):
            cell_val = ws.cell(row=2, column=col).value or ws.cell(row=1, column=col).value
            parsed_d = parse_cell_date(cell_val, target_year)
            if parsed_d and parsed_d == target_date:
                date_col = col
                break
        if not date_col: continue

        for row in range(3, ws.max_row + 1):
            name_val = ws.cell(row=row, column=2).value
            if not name_val: continue
            name_str = str(name_val).strip()
            if name_str in ("구분", "월통계", "총인원") or "통계" in name_str: break
                
            c_val = ws.cell(row=row, column=date_col).value
            is_present = False
            if c_val is True: is_present = True
            elif isinstance(c_val, str):
                if c_val.strip().upper() in ("TRUE", "O", "○", "ㅇ", "출석", "참석", "Y", "YES", "1", "●"):
                    is_present = True
            elif isinstance(c_val, (int, float)) and c_val == 1: is_present = True

            if is_present: people[part_code].append(name_str)

    return people

def row_targets(total):
    caps = [18, 18, 17, 16]
    target_total = min(max(1, int(total)), sum(caps))
    candidates = []
    for r_equal in range(0, min(caps[0], caps[1]) + 1):
        r1 = r2 = r_equal
        for r3 in range(0, caps[2] + 1):
            if r2 < r3: continue
            r4 = target_total - r1 - r2 - r3
            if 0 <= r4 <= caps[3]:
                if r3 < r4: continue
                candidates.append(((r2 - r3) * 100 + (r3 - r4) * 50, [r1, r2, r3, r4]))
    if not candidates:
        for r_equal in range(min(caps[0], caps[1]), -1, -1):
            r1 = r2 = r_equal
            rem = target_total - r1 - r2
            if rem >= 0:
                r3 = min(rem, caps[2])
                r4 = rem - r3
                if r4 <= caps[3]:
                    candidates.append((0, [r1, r2, r3, r4]))
                    break
    return min(candidates, key=lambda x: x[0])[1]

def allocate(people_dict, total_seats):
    people = {p: list(people_dict[p]) for p in ("S", "A", "T", "B")}
    rt = row_targets(total_seats)
    
    male_count = len(people["T"]) + len(people["B"])
    organ_row = 3 if male_count < rt[3] or (rt[2] < rt[3]) else 2
    
    caps = rt[:]
    if caps[organ_row] > 0: caps[organ_row] -= 1

    leftover = {"S": [], "A": [], "T": [], "B": []}
    total_capacity = sum(caps)
    total_attending = sum(len(people[p]) for p in ("S", "A", "T", "B"))
    
    # 성도석 이동 처리 (알토 파트)
    if total_attending > total_capacity:
        excess_count = total_attending - total_capacity
        alto_out_candidates = [m for m in ALTO_OUT_PRIORITY if m in people["A"]]
        for m in people["A"]:
            if m not in alto_out_candidates: alto_out_candidates.append(m)
        moved_alto = []
        while excess_count > 0 and alto_out_candidates:
            m = alto_out_candidates.pop(0)
            moved_alto.append(m)
            people["A"].remove(m)
            excess_count -= 1
        leftover["A"] = moved_alto

    rows = [[], [], [], []]

    s_row3_fixed = [m for m in SOPRANO_ROW3_PRIORITY if m in people["S"]]
    for m in s_row3_fixed: people["S"].remove(m)

    female_a = [m for m in BASE_SEAT_ORDER[1] + BASE_SEAT_ORDER[2] if m in people["A"]]
    for m in people["A"]:
        if m not in female_a: female_a.append(m)
        
    female_s = [m for m in BASE_SEAT_ORDER[1] + BASE_SEAT_ORDER[2] if m in people["S"]]
    for m in people["S"]:
        if m not in female_s: female_s.append(m)

    max_per_row = caps[0]
    total_females = len(female_a) + len(female_s)
    front_per_row = min(max_per_row, total_females // 2)

    count_a_per_row = min(len(female_a) // 2, front_per_row)
    
    rows[0] = [("A", m) for m in female_a[:count_a_per_row]]
    rows[1] = [("A", m) for m in female_a[count_a_per_row:count_a_per_row * 2]]
    
    unplaced_a = female_a[count_a_per_row * 2:]
    need_s = front_per_row - len(rows[0])
    rows[0] += [("S", m) for m in female_s[:need_s]]
    rows[1] += [("S", m) for m in female_s[need_s:need_s * 2]]
    
    people["A"], people["S"] = unplaced_a, female_s[need_s * 2:]

    # 3열 배정
    for name in s_row3_fixed:
        if len(rows[2]) >= caps[2]: break
        rows[2].append(("S", name))

    for name in BASE_SEAT_ORDER[3]:
        if len(rows[2]) >= caps[2]: break
        if name in s_row3_fixed: continue
        for p in ("S", "A", "T", "B"):
            if name in people[p]:
                rows[2].append((p, name))
                people[p].remove(name)
                break

    while len(rows[2]) < caps[2]:
        cand = [p for p in ("S", "A", "T", "B") if people[p]]
        if not cand: break
        rows[2].append((cand[0], people[cand[0]].pop(0)))

    # 4열 배정
    for name in BASE_SEAT_ORDER[4]:
        if len(rows[3]) >= caps[3]: break
        for p in ("T", "B"):
            if name in people[p]:
                rows[3].append((p, name))
                people[p].remove(name)
                break

    while len(rows[3]) < caps[3]:
        cand = [p for p in ("T", "B") if people[p]]
        if not cand: break
        rows[3].append((cand[0], people[cand[0]].pop(0)))

    # 파트 순서 정렬 (B -> T -> A -> S -> ORG) 및 파트 내 역순 정렬
    reversed_part_order = ["B", "T", "A", "S", "ORG"]
    
    final_rows, row_part_counts = [], []
    for r_idx in range(4):
        curr_row = list(rows[r_idx])
        if r_idx == organ_row:
            curr_row.append(("ORG", "오르간"))
        
        part_grouped = {p: [] for p in reversed_part_order}
        for part, name in curr_row:
            part_grouped[part].append(name)
        
        sorted_row = []
        for p in reversed_part_order:
            members = list(reversed(part_grouped[p]))
            for m in members:
                sorted_row.append((p, m))

        final_rows.append(sorted_row)
        p_counts = {p: len(part_grouped[p]) for p in ("S", "A", "T", "B", "ORG")}
        row_part_counts.append(p_counts)

    return final_rows, leftover, rt, organ_row, row_part_counts

class RequestModel(BaseModel):
    date_str: str
    total_seats: int = 67

@app.post("/api/allocate")
def run_allocation(req: RequestModel):
    try:
        target = dt.date.fromisoformat(req.date_str)
        service = google_drive_service()
        file_info = find_drive_file(service)
        
        file_stream = download_excel_to_stream(service, file_info["id"])
        people = read_attendance(file_stream, target)
        
        attending_by_part = {p: len(people[p]) for p in ("S", "A", "T", "B")}
        rows, leftover, rt, org, row_part_counts = allocate(people, req.total_seats)
        
        return {
            "status": "success",
            "date": str(target),
            "attending_count": sum(attending_by_part.values()),
            "attending_by_part": attending_by_part,
            "row_targets": rt,
            "organ_row": org + 1,
            "rows": rows,
            "row_part_counts": row_part_counts,
            "sungdoseok": leftover  # 미배치 인원 대신 성도석 명단 반환
        }
    except Exception as e:
        logger.error(f"오류 발생: {str(e)}")
        raise HTTPException(status_code=400, detail=str(e))
2. 프론트엔드 연동 가이드 (index.html 내 JS)
웹 화면(프론트엔드)에서 4가지 요청 사항대로 표시되도록 렌더링 부분을 아래 스크립트로 작성해주세요.

JavaScript
// API 응답 데이터를 화면에 그리거나 표출하는 함수 예시
function renderResult(data) {
  // 3. 출석인원 옆에 각 파트 출석인원 표시 예시 (ex: 출석인원 : 총 70명 (S 24, A 19, T 12, B 15))
  const p = data.attending_by_part;
  const attendanceText = `출석인원 : 총 ${data.attending_count}명 (S ${p.S || 0}, A ${p.A || 0}, T ${p.T || 0}, B ${p.B || 0})`;
  document.getElementById("attendance-summary").innerText = attendanceText;

  // 좌석배치 표출 영역
  const container = document.getElementById("seat-map-container");
  container.innerHTML = "";

  data.rows.forEach((row, rIdx) => {
    const rowNum = rIdx + 1;
    const target = data.row_targets[rIdx];
    const counts = data.row_part_counts[rIdx];

    // 2. 각 열마다 각 파트 인원 합계를 표시 예시 (ex : [1열] A 7명, S 10명 , 목표 17 명)
    const partParts = [];
    if (counts.A > 0) partParts.push(`A ${counts.A}명`);
    if (counts.S > 0) partParts.push(`S ${counts.S}명`);
    if (counts.T > 0) partParts.push(`T ${counts.T}명`);
    if (counts.B > 0) partParts.push(`B ${counts.B}명`);
    if (counts.ORG > 0) partParts.push(`오르간 1명`);

    const headerText = `[${rowNum}열] ${partParts.join(', ')} , 목표 ${target} 명`;
    
    // 열 헤더 생성
    const rowHeader = document.createElement("h3");
    rowHeader.className = "row-header";
    rowHeader.innerText = headerText;
    container.appendChild(rowHeader);

    // 1. 각 파트 이름에 원래대로 색깔을 주고 이름 앞에 [] 표시 대신 , 콤마로 구분
    const rowDiv = document.createElement("div");
    rowDiv.className = "row-container";

    const nameSpans = row.map(([partCode, name]) => {
      // 파트 코드별 CSS 클래스 (색상 부여)
      const colorClass = `part-${partCode.toLowerCase()}`;
      return `<span class="seat-item ${colorClass}">${name}</span>`;
    });

    // [] 텍스트 표기 없이 색상 스팬을 ,(콤마)로 구분하여 연결
    rowDiv.innerHTML = nameSpans.join(", ");
    container.appendChild(rowDiv);
  });

  // 4. 미배치 인원 문구 대신 '성도석'이란 문구로 대체
  const sungdoseokContainer = document.getElementById("sungdoseok-container");
  const sungdoList = [];
  Object.values(data.sungdoseok).forEach(arr => sungdoList.push(...arr));

  if (sungdoList.length > 0) {
    sungdoseokContainer.innerHTML = `<h4>성도석 : ${sungdoList.join(", ")}</h4>`;
  } else {
    sungdoseokContainer.innerHTML = `<h4>성도석 : 없음</h4>`;
  }
}